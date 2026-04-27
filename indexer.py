#!/usr/bin/env python3
"""
CBW RAG Indexer - Scans files, extracts text, chunks, embeds via Ollama, stores in pgvector.
"""

import os, sys, json, hashlib, argparse, mimetypes, subprocess, time
from pathlib import Path
from datetime import datetime, timezone
from typing import List, Dict, Optional, Tuple, Generator
from dataclasses import dataclass

import psycopg2
import ollama
import tiktoken
import chardet
import magic
import os
import json
from concurrent.futures import ThreadPoolExecutor, as_completed
from psycopg2.extras import execute_values
from pgvector.psycopg2 import register_vector

DB_DSN = os.environ.get("CBW_RAG_DATABASE", "postgresql://cbwinslow:123qweasd@localhost:5432/cbw_rag")
OLLAMA_URL = os.environ.get("CBW_RAG_OLLAMA_URL", "http://localhost:11434")
EMBEDDING_MODEL = os.environ.get("CBW_RAG_EMBEDDING_MODEL", "nomic-embed-text")

# Increased chunk size for fewer database rows and faster retrieval
CHUNK_SIZE = 1024
CHUNK_OVERLAP = 128
MAX_FILE_SIZE = 50 * 1024 * 1024

TEXT_EXTS = {
    '.txt','.md','.markdown','.rst','.log','.cfg','.conf','.ini',
    '.yaml','.yml','.toml','.json','.xml','.csv','.tsv',
    '.html','.htm','.css','.js','.ts','.jsx','.tsx',
    '.py','.pyi','.rb','.rs','.go','.java','.kt','.kts',
    '.c','.h','.cpp','.hpp','.cc','.cxx','.cs',
    '.sh','.bash','.zsh','.fish','.ps1','.bat','.cmd',
    '.sql','.r','.jl','.lua','.pl','.pm','.php','.swift',
    '.dockerfile','.env','.gitignore','.editorconfig',
    '.tf','.hcl','.proto','.graphql','.gql',
    '.makefile','.cmake','.gradle','.bib','.tex',
}

SKIP_DIRS = {
    '.git','.svn','.hg','node_modules','__pycache__','.pytest_cache',
    '.mypy_cache','venv','.venv','env','.env','.tox','.eggs','.cache',
    '.cargo','.rustup','.local','.npm','.nvm','.bun','vendor','target',
    'build','dist','.next','.opencode','.kilo','.windsurf-server',
    'site-packages','.bun','.rustup','.cache','__pycache__',
}

BINARY_EXTS = {
    '.png','.jpg','.jpeg','.gif','.bmp','.ico','.svg','.webp',
    '.mp3','.mp4','.avi','.mkv','.mov','.wav','.flac','.ogg',
    '.zip','.tar','.gz','.bz2','.xz','.7z','.rar','.tgz',
    '.exe','.dll','.so','.dylib','.o','.a','.lib',
    '.pyc','.pyo','.class','.beam','.db','.sqlite','.sqlite3',
    '.woff','.woff2','.ttf','.otf','.eot',
}

enc = tiktoken.get_encoding("cl100k_base")

def count_tokens(text):
    return len(enc.encode(text))

def classify_file(path):
    ext = path.suffix.lower()
    name = path.name.lower()
    lang_map = {
        '.py':'python','.pyi':'python','.js':'javascript','.jsx':'javascript',
        '.ts':'typescript','.tsx':'typescript','.rb':'ruby','.rs':'rust',
        '.go':'go','.java':'java','.kt':'kotlin','.c':'c','.h':'c',
        '.cpp':'cpp','.hpp':'cpp','.cs':'csharp','.sh':'shell','.bash':'bash',
        '.zsh':'zsh','.sql':'sql','.r':'r','.jl':'julia','.lua':'lua',
        '.php':'php','.swift':'swift','.html':'html','.htm':'html','.css':'css',
        '.md':'markdown','.yaml':'yaml','.yml':'yaml','.json':'json',
        '.xml':'xml','.toml':'toml','.tf':'terraform','.hcl':'hcl',
    }
    cat_map = {
        '.py':'code','.js':'code','.ts':'code','.jsx':'code','.tsx':'code',
        '.rb':'code','.rs':'code','.go':'code','.java':'code','.kt':'code',
        '.c':'code','.h':'code','.cpp':'code','.hpp':'code','.cs':'code',
        '.sh':'code','.bash':'code','.zsh':'code',
        '.yaml':'config','.yml':'config','.toml':'config','.json':'config',
        '.xml':'config','.conf':'config','.cfg':'config','.ini':'config','.env':'config',
    }
    lang = lang_map.get(ext)
    if name in ('makefile','dockerfile'):
        lang = name
    cat = cat_map.get(ext, 'document')
    return cat, lang

def get_git_info(path):
    try:
        r = subprocess.run(['git','-C',str(path.parent),'rev-parse','--show-toplevel'], capture_output=True, text=True, timeout=5)
        if r.returncode != 0: return None, None, None, None
        repo = r.stdout.strip()
        branch = subprocess.run(['git','-C',str(path.parent),'branch','--show-current'], capture_output=True, text=True, timeout=5).stdout.strip() or None
        commit = subprocess.run(['git','-C',str(path.parent),'log','-1','--format=%H','--',str(path)], capture_output=True, text=True, timeout=5).stdout.strip() or None
        st = subprocess.run(['git','-C',str(path.parent),'status','--porcelain',str(path)], capture_output=True, text=True, timeout=5)
        gs = 'modified' if st.stdout.strip() else ('tracked' if commit else 'untracked')
        return repo, branch, commit, gs
    except:
        return None, None, None, None

def extract_metadata(path):
    try:
        stat = path.stat(follow_symlinks=False)
    except (OSError, PermissionError):
        return None

    ext = path.suffix.lower()
    if ext in BINARY_EXTS:
        return None

    mime_type, _ = mimetypes.guess_type(str(path))
    if not mime_type:
        try: mime_type = magic.from_file(str(path), mime=True)
        except: mime_type = None

    encoding = None
    try:
        with open(path, 'rb') as f: raw = f.read(4096)
        detected = chardet.detect(raw)
        encoding = detected.get('encoding')
    except: pass

    try:
        import pwd, grp
        owner = pwd.getpwuid(stat.st_uid).pw_name
        group = grp.getgrgid(stat.st_gid).gr_name
    except:
        owner, group = str(stat.st_uid), str(stat.st_gid)

    cat, lang = classify_file(path)

    line_count = None
    if cat in ('code','document','config'):
        try:
            with open(path, 'rb') as f:
                line_count = sum(1 for _ in f)
        except: pass

    try:
        with open(path, 'rb') as f:
            content_hash = hashlib.sha256(f.read()).hexdigest()
    except:
        return None

    git_repo, git_branch, git_commit, git_status = get_git_info(path)

    return {
        'source_path': str(path.resolve()),
        'file_name': path.name,
        'file_extension': ext,
        'file_size': stat.st_size,
        'file_mode': stat.st_mode,
        'file_owner': owner,
        'file_group': group,
        'mime_type': mime_type,
        'file_category': cat,
        'detected_language': lang,
        'encoding': encoding,
        'line_count': line_count,
        'file_content_hash': content_hash,
        'file_created_at': datetime.fromtimestamp(stat.st_ctime, tz=timezone.utc),
        'file_modified_at': datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc),
        'git_repo_root': git_repo,
        'git_branch': git_branch,
        'git_last_commit': git_commit,
        'git_status': git_status,
    }

def scan_directory(root, skip_dirs=SKIP_DIRS):
    root_path = Path(root).expanduser().resolve()
    for dirpath, dirnames, filenames in os.walk(root_path, followlinks=False):
        dirnames[:] = [d for d in dirnames if d not in skip_dirs and not d.startswith('.cache')]
        for fname in filenames:
            p = Path(dirpath) / fname
            if p.is_symlink() or not p.is_file():
                continue
            if p.suffix.lower() in BINARY_EXTS:
                continue
            yield p

def extract_text(path, category):
    ext = path.suffix.lower()
    if ext == '.pdf':
        try:
            r = subprocess.run(['pdftotext', str(path), '-'], capture_output=True, text=True, timeout=60)
            if r.returncode == 0: return r.stdout
        except: pass
        return None
    if ext == '.docx':
        try:
            from docx import Document
            return '\n'.join(p.text for p in Document(str(path)).paragraphs if p.text.strip())
        except: pass
        return None
    if ext in ('.xlsx', '.xls'):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(path), read_only=True, data_only=True)
            lines = []
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    line = '\t'.join(str(c) if c is not None else '' for c in row)
                    if line.strip(): lines.append(line)
            wb.close()
            return '\n'.join(lines)
        except: pass
        return None
    if category in ('code','document','config','unknown'):
        try:
            with open(path, 'rb') as f: raw = f.read(MAX_FILE_SIZE)
            detected = chardet.detect(raw)
            enc_name = detected.get('encoding', 'utf-8') or 'utf-8'
            return raw.decode(enc_name, errors='replace')
        except: pass
    return None

def chunk_text(text, chunk_size=CHUNK_SIZE, overlap=CHUNK_OVERLAP):
    if not text or not text.strip(): return []
    lines = text.split('\n')
    chunks = []
    current = []
    current_tokens = 0
    start_line = 1

    for i, line in enumerate(lines, 1):
        lt = count_tokens(line)
        if current_tokens + lt > chunk_size and current:
            content = '\n'.join(current)
            if content.strip():
                chunks.append({
                    'content': content,
                    'content_hash': hashlib.sha256(content.encode()).hexdigest(),
                    'token_count': current_tokens,
                    'char_count': len(content),
                    'start_line': start_line,
                    'end_line': i - 1,
                })
            overlap_lines = []
            overlap_t = 0
            for prev in reversed(current):
                pl = count_tokens(prev)
                if overlap_t + pl > overlap: break
                overlap_lines.insert(0, prev)
                overlap_t += pl
            current = overlap_lines + [line]
            current_tokens = overlap_t + lt
            start_line = i - len(overlap_lines)
        else:
            current.append(line)
            current_tokens += lt

    if current:
        content = '\n'.join(current)
        if content.strip():
            chunks.append({
                'content': content,
                'content_hash': hashlib.sha256(content.encode()).hexdigest(),
                'token_count': current_tokens,
                'char_count': len(content),
                'start_line': start_line,
                'end_line': len(lines),
            })
    return chunks

def generate_embeddings(texts, model=EMBEDDING_MODEL):
    """Generate embeddings in larger batches (default 64) using Ollama's batch API.
    Falls back to per‑item calls if the batch request fails.
    """
    embeddings = []
    batch_size = min(64, len(texts))
    for i in range(0, len(texts), batch_size):
        batch = [t[:2000] for t in texts[i:i+batch_size]]
        try:
            r = ollama.embed(model=model, input=batch)
            embeddings.extend(r.get('embeddings', []))
        except Exception as e:
            print(f"  Embedding batch error: {e}", file=sys.stderr)
            for txt in batch:
                try:
                    r2 = ollama.embeddings(model=model, prompt=txt[:2000])
                    embeddings.append(r2.get('embedding', [0.0] * 768))
                except Exception:
                    embeddings.append([0.0] * 768)
    return embeddings

def store_file(cur, meta):
    """Upsert file, return (file_id, needs_indexing)."""
    cur.execute("SELECT id, content_hash FROM files WHERE file_path = %s", (meta['source_path'],))
    row = cur.fetchone()
    if row:
        fid, old_hash = row
        if old_hash == meta['file_content_hash']:
            return fid, False
        cur.execute("DELETE FROM file_chunks WHERE file_id = %s", (fid,))
        cur.execute("""UPDATE files SET file_name=%s, file_extension=%s, file_size_bytes=%s,
            mime_type=%s, content_hash=%s, modified_at=%s, metadata=%s, is_deleted=false
            WHERE id=%s""",
            (meta['file_name'], meta['file_extension'], meta['file_size'],
             meta['mime_type'], meta['file_content_hash'], meta['file_modified_at'],
             json.dumps({'file_category': meta.get('file_category'), 'detected_language': meta.get('detected_language'),
                        'encoding': meta.get('encoding'), 'line_count': meta.get('line_count'),
                        'file_mode': meta.get('file_mode'), 'file_owner': meta.get('file_owner'),
                        'file_group': meta.get('file_group'), 'git_repo_root': meta.get('git_repo_root'),
                        'git_branch': meta.get('git_branch'), 'git_last_commit': meta.get('git_last_commit'),
                        'git_status': meta.get('git_status')}), fid))
        return fid, True
    else:
        cur.execute("""INSERT INTO files (file_path,file_name,file_extension,file_size_bytes,
            mime_type,content_hash,created_at,modified_at,metadata)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
            (meta['source_path'], meta['file_name'], meta['file_extension'], meta['file_size'],
             meta['mime_type'], meta['file_content_hash'], meta['file_created_at'], meta['file_modified_at'],
             json.dumps({'file_category': meta.get('file_category'), 'detected_language': meta.get('detected_language'),
                        'encoding': meta.get('encoding'), 'line_count': meta.get('line_count'),
                        'file_mode': meta.get('file_mode'), 'file_owner': meta.get('file_owner'),
                        'file_group': meta.get('file_group'), 'git_repo_root': meta.get('git_repo_root'),
                        'git_branch': meta.get('git_branch'), 'git_last_commit': meta.get('git_last_commit'),
                        'git_status': meta.get('git_status')})))
        return cur.fetchone()[0], True

def store_chunks(cur, file_id, chunks, embeddings, model):
    """Bulk insert chunks for a single file using execute_values for efficiency."""
    from psycopg2.extras import execute_values
    rows = []
    for i, chunk in enumerate(chunks):
        emb = embeddings[i] if i < len(embeddings) else [0.0] * 768
        # Pass vector as a Python list; psycopg2 will convert to pgvector
        rows.append((
            file_id,
            i,
            chunk['content'],
            chunk['content_hash'],
            emb,  # vector
            chunk['token_count'],
            chunk['char_count'],
            chunk['start_line'],
            chunk['end_line'],
            json.dumps({'model': model})
        ))
    sql = """
        INSERT INTO file_chunks (
            file_id, chunk_index, chunk_text, chunk_hash,
            embedding, token_count, char_count, line_start, line_end, metadata
        ) VALUES %s
        ON CONFLICT (file_id, chunk_index) DO UPDATE SET
            chunk_text = EXCLUDED.chunk_text,
            chunk_hash = EXCLUDED.chunk_hash,
            embedding = EXCLUDED.embedding,
            token_count = EXCLUDED.token_count,
            char_count = EXCLUDED.char_count,
            line_start = EXCLUDED.line_start,
            line_end = EXCLUDED.line_end,
            metadata = EXCLUDED.metadata,
            modified_at = NOW()
    """
    execute_values(cur, sql, rows, template=None, page_size=100)
    cur.execute("UPDATE files SET indexed_at=NOW() WHERE id=%s", (file_id,))

def index_files(paths, dry_run=False, resume=False, batch_size=25, verbose=False, model=EMBEDDING_MODEL):
    conn = psycopg2.connect(DB_DSN)
    conn.autocommit = False
    # Enable automatic conversion for pgvector types
    register_vector(conn)
    cur = conn.cursor()

    stats = {'scanned':0,'indexed':0,'unchanged':0,'skipped':0,'errors':0,'chunks':0,'embeddings':0}
    batch = []

    try:
        for root in paths:
            print(f"\nScanning: {root}")
            for path in scan_directory(root):
                stats['scanned'] += 1
                if stats['scanned'] % 100 == 0:
                    print(f"  {stats['scanned']} scanned | {stats['indexed']} indexed | {stats['unchanged']} unchanged | {stats['skipped']} skipped")

                meta = extract_metadata(path)
                if meta is None:
                    stats['skipped'] += 1
                    continue

                if meta['file_size'] > MAX_FILE_SIZE:
                    stats['skipped'] += 1
                    continue

                if dry_run:
                    print(f"  WOULD INDEX: {meta['source_path']} [{meta['file_category']}/{meta['detected_language']}]")
                    continue

                fid, needs_idx = store_file(cur, meta)
                if not needs_idx:
                    stats['unchanged'] += 1
                    conn.rollback()
                    continue

                text = extract_text(path, meta['file_category'])
                if not text or not text.strip():
                    stats['skipped'] += 1
                    conn.rollback()
                    continue

                chunks = chunk_text(text)
                if not chunks:
                    stats['skipped'] += 1
                    conn.rollback()
                    continue

                batch.append((fid, chunks, meta['source_path']))

                if len(batch) >= batch_size:
                    _process_batch(cur, conn, batch, model, stats, verbose)
                    batch = []
                    conn.commit()

            if batch:
                _process_batch(cur, conn, batch, model, stats, verbose)
                batch = []
                conn.commit()

    finally:
        cur.close()
        conn.close()

    print(f"\n--- Done ---")
    print(f"Scanned: {stats['scanned']} | Indexed: {stats['indexed']} | Unchanged: {stats['unchanged']} | Skipped: {stats['skipped']} | Errors: {stats['errors']}")
    print(f"Chunks: {stats['chunks']} | Embeddings: {stats['embeddings']}")
    return stats

def _process_batch(cur, conn, batch, model, stats, verbose):
    all_texts = []
    mapping = []
    for fid, chunks, path in batch:
        for i, ch in enumerate(chunks):
            mapping.append((fid, i, len(all_texts)))
            all_texts.append(ch['content'])

    if not all_texts: return
    if verbose: print(f"  Embedding {len(all_texts)} chunks...")
    embeddings = generate_embeddings(all_texts, model)

    emb_by_file = {}
    for fid, ci, ti in mapping:
        emb_by_file.setdefault(fid, []).append(embeddings[ti])

    for fid, chunks, path in batch:
        try:
            store_chunks(cur, fid, chunks, emb_by_file.get(fid, []), model)
            stats['indexed'] += 1
            stats['chunks'] += len(chunks)
            stats['embeddings'] += len(chunks)
            if verbose: print(f"  OK: {path} ({len(chunks)} chunks)")
        except Exception as e:
            stats['errors'] += 1
            if verbose: print(f"  ERR: {path}: {e}")
            conn.rollback()

def main():
    parser = argparse.ArgumentParser(description='CBW RAG Indexer')
    parser.add_argument('paths', nargs='+', help='Directories to index')
    parser.add_argument('--dry-run', action='store_true')
    parser.add_argument('--resume', action='store_true')
    parser.add_argument('--batch-size', type=int, default=50)
    parser.add_argument('--verbose', '-v', action='store_true')
    parser.add_argument('--model', default=EMBEDDING_MODEL)
    args = parser.parse_args()

    index_files(args.paths, dry_run=args.dry_run, resume=args.resume,
                batch_size=args.batch_size, verbose=args.verbose, model=args.model)

if __name__ == '__main__':
    main()
